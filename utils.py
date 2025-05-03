import os
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from PIL import Image
from datetime import datetime
import io
import pandas as pd
import streamlit as st
import requests
import base64


def load_excel_data(path):
    """Load Excel file from path or GitHub"""
    try:
        df = pd.read_excel(path, sheet_name="Sheet1")
        wb = openpyxl.load_workbook(path)
        return df, wb
    except Exception as e:
        st.error(f"Error loading Excel: {str(e)}")
        raise


def get_task_ids(df):
    return df["Task ID"].dropna().astype(str).tolist()


def insert_image(ws, img_bytes, row):
    img = Image.open(img_bytes)
    img.thumbnail((600, 400))
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    img_obj = OpenpyxlImage(bio)
    cell = f"A{row}"
    ws.add_image(img_obj, cell)
    ws.column_dimensions['A'].width = 60
    ws.row_dimensions[row].height = 100
    return row + 15


def upload_to_github(file_path, repo_owner, repo_name, token, file_name):
    """Upload the updated Excel file to GitHub."""
    url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{file_name}"

    with open(file_path, "rb") as file:
        content = base64.b64encode(file.read()).decode("utf-8")

    headers = {
        "Authorization": f"token {token}",
        "Content-Type": "application/json",
    }

    # Create commit data
    data = {
        "message": "Update test results",
        "content": content,
        "branch": "main"  # Use your default branch
    }

    response = requests.put(url, headers=headers, json=data)
    if response.status_code == 201:
        print(f"Successfully uploaded {file_name} to GitHub")
    else:
        print(f"Failed to upload to GitHub: {response.status_code} - {response.text}")


def save_screenshots_to_excel(excel_path, df_main, wb, task_id, tester_name, test_result, comment, screenshots):
    if isinstance(excel_path, (str, os.PathLike)):
        def normalize_id(tid):
            return str(int(float(tid))) if float(tid).is_integer() else str(tid)

        normalized_task_id = normalize_id(task_id)
        task_info = df_main[df_main["Task ID"].astype(str).apply(normalize_id) == normalized_task_id].iloc[0]
        main_task_id = str(task_id).split('.')[0]
        sheet_name = f"Task ID {main_task_id}"

        new_sheet = False
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            current_row = 1
            new_sheet = True
        else:
            ws = wb[sheet_name]
            current_row = None

            search_label = "Task" if '.' not in str(task_id) else "Subtask"
            search_text = f"Task {task_id}"

            for row in ws.iter_rows(min_row=1, max_col=2):
                if row[0].value == search_label and row[1].value == search_text:
                    current_row = row[0].row
                    while True:
                        current_row += 1
                        next_label = ws.cell(row=current_row, column=1).value
                        if next_label in [None, "", "Task", "Subtask"]:
                            break
                    break

            if current_row is None:
                current_row = ws.max_row + 2

        def write_row(label, value, bold=False):
            nonlocal current_row
            font_style = Font(bold=bold)
            ws.cell(row=current_row, column=1, value=label).font = font_style
            ws.cell(row=current_row, column=2, value=value).font = font_style
            current_row += 1

        is_new_block = new_sheet or current_row == ws.max_row + 2

        if is_new_block:
            label = "Task" if '.' not in str(task_id) else "Subtask"
            write_row(label, f"Task {task_id}", bold=True)
            write_row("Navigation", task_info["Navigation"], bold=True)
            write_row("Tester Name", tester_name, bold=True)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            write_row("Timestamp", timestamp, bold=True)

        for screenshot in screenshots:
            current_row = insert_image(ws, screenshot, current_row)

        result_row = current_row
        write_row("Test Result", test_result, bold=True)

        fill_color = {
            "Pass": "90EE90",
            "Fail": "FF6347",
            "Hold": "FFB6C1"
        }.get(test_result, "FFFFFF")

        result_cell = ws.cell(row=result_row, column=2)
        result_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        if comment:
            write_row("Comment", comment, bold=True)

        current_row += 2

        main_ws = wb["Sheet1"]
        for row in range(2, main_ws.max_row + 1):
            cell_task_id = str(main_ws.cell(row=row, column=1).value)
            if normalize_id(cell_task_id) == normalized_task_id:
                main_ws.cell(row=row, column=5).value = tester_name
                main_ws.cell(row=row, column=6).value = test_result
                main_ws.cell(row=row, column=7).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                result_fill_color = {
                    "Pass": "90EE90",
                    "Fail": "FF6347",
                    "Hold": "FFB6C1"
                }.get(test_result, "FFFFFF")
                result_cell = main_ws.cell(row=row, column=6)
                result_cell.fill = PatternFill(start_color=result_fill_color, end_color=result_fill_color, fill_type="solid")
                break

        def update_summary_sheet():
            summary_sheet_name = "Summary"
            if summary_sheet_name not in wb.sheetnames:
                summary_ws = wb.create_sheet(summary_sheet_name)
            else:
                summary_ws = wb[summary_sheet_name]

            df_updated = pd.DataFrame(main_ws.values)
            headers = df_updated.iloc[0].tolist()
            df_updated.columns = [str(col).strip() if col is not None else f"Column_{i}" for i, col in enumerate(headers)]
            df_updated = df_updated.drop(index=0).reset_index(drop=True)

            total_tasks = df_updated.shape[0]
            pass_count = df_updated[df_updated["Test Result"] == "Pass"].shape[0]
            fail_count = df_updated[df_updated["Test Result"] == "Fail"].shape[0]
            hold_count = df_updated[df_updated["Test Result"] == "Hold"].shape[0]
            pass_rate = f"{(pass_count / total_tasks * 100):.2f}%" if total_tasks else "0%"

            summary_data = [
                ("Total Tasks", total_tasks),
                ("Pass", pass_count),
                ("Fail", fail_count),
                ("Hold", hold_count),
                ("Pass Rate", pass_rate),
                ("Last Updated Task ID", task_id),
                ("Last Updated By", tester_name),
                ("Last Updated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ]

            for i, (label, value) in enumerate(summary_data, start=1):
                summary_ws.cell(row=i, column=1).value = label
                summary_ws.cell(row=i, column=2).value = value
                summary_ws.cell(row=i, column=1).font = Font(bold=True)

            while summary_ws._charts:
                summary_ws._charts.pop()

            labels = Reference(summary_ws, min_col=1, min_row=2, max_row=4)
            data = Reference(summary_ws, min_col=2, min_row=2, max_row=4)
            pie_chart = PieChart()
            pie_chart.title = "Test Result Summary"
            pie_chart.add_data(data, titles_from_data=False)
            pie_chart.set_categories(labels)
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showVal = True
            summary_ws.add_chart(pie_chart, "D2")

            progress_row = 12
            percent_complete = (pass_count + fail_count + hold_count) / total_tasks if total_tasks else 0
            filled_blocks = int(percent_complete * 20)
            empty_blocks = 20 - filled_blocks
            progress_bar = filled_blocks * "█" + empty_blocks * "-"
            summary_ws.cell(row=progress_row, column=1).value = "Progress"
            summary_ws.cell(row=progress_row, column=2).value = f"[{progress_bar}] {int(percent_complete * 100)}%"
            summary_ws.cell(row=progress_row, column=2).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                    # 1st graph: Task Completion Over Time
            if "Timestamp" in df_updated.columns:
                df_filtered = df_updated[df_updated["Timestamp"].notna()]
                df_filtered["Date"] = pd.to_datetime(df_filtered["Timestamp"]).dt.date
                date_summary = df_filtered.groupby("Date").size().reset_index(name="Tasks Completed")

                # Add headers before inserting date_summary
                date_header_row = 19
                summary_ws.cell(row=date_header_row, column=1).value = "Date"
                summary_ws.cell(row=date_header_row, column=2).value = "Test Count"
                summary_ws.cell(row=date_header_row, column=1).font = Font(bold=True)
                summary_ws.cell(row=date_header_row, column=2).font = Font(bold=True)

                for i, (date, count) in enumerate(date_summary.itertuples(index=False), start=20):
                    summary_ws.cell(row=i, column=1).value = str(date)
                    summary_ws.cell(row=i, column=2).value = count

                line_chart = LineChart()
                line_chart.title = "Task Completion Over Time"
                line_chart.y_axis.title = "Tasks Completed"
                line_chart.x_axis.title = "Date"
                line_chart.add_data(Reference(summary_ws, min_col=2, min_row=19, max_row=19+len(date_summary)), titles_from_data=True)
                line_chart.set_categories(Reference(summary_ws, min_col=1, min_row=20, max_row=19+len(date_summary)))
                summary_ws.add_chart(line_chart, "D18")

            ## 2nd graph: Tasks Completed Per Tester
            if "Tester Name" in df_updated.columns and "Test Result" in df_updated.columns:
                df_filtered = df_updated[df_updated["Test Result"].notna()]
                tester_summary = df_filtered["Tester Name"].value_counts().reset_index()
                tester_summary.columns = ["Tester Name", "Tasks Completed"]

                # Add headers before inserting tester_summary
                tester_header_row = 39
                summary_ws.cell(row=tester_header_row, column=1).value = "Tester Name"
                summary_ws.cell(row=tester_header_row, column=2).value = "Test Count"
                summary_ws.cell(row=tester_header_row, column=1).font = Font(bold=True)
                summary_ws.cell(row=tester_header_row, column=2).font = Font(bold=True)

                for i, (name, count) in enumerate(tester_summary.itertuples(index=False), start=40):
                    summary_ws.cell(row=i, column=1).value = name
                    summary_ws.cell(row=i, column=2).value = count

                bar_chart = BarChart()
                bar_chart.title = "Tasks Completed Per Tester"
                bar_chart.y_axis.title = "Task Count"
                bar_chart.x_axis.title = "Tester"
                bar_chart.add_data(Reference(summary_ws, min_col=2, min_row=39, max_row=39+len(tester_summary)), titles_from_data=True)
                bar_chart.set_categories(Reference(summary_ws, min_col=1, min_row=40, max_row=39+len(tester_summary)))
                summary_ws.add_chart(bar_chart, "D35")


        update_summary_sheet()

        # Save locally
        
        wb.save(excel_path)

# Upload to GitHub
        upload_to_github(
            excel_path,
            "Ai-TestingApp",                # 🔁 Replace with your GitHub username
            "Ai-Testing-Tool",                      # 🔁 Replace with your repository name
            st.secrets["GITHUB_TOKEN"],            # ✅ Keep this if your token is stored in .streamlit/secrets.toml
            "main/main_excel.xlsx"         # 🔁 Replace with actual repo file path
        )


        # Upload to GitHub
        #upload_to_github(excel_path, "Ai-TestingApp", "Ai-Testing-Tool", st.secrets["GITHUB_TOKEN"], excel_path)
def upload_to_github(excel_path, repo_owner, repo_name, token, repo_file_path):
    """Uploads the updated Excel file back to GitHub."""
    import base64
    import requests

    api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{repo_file_path}"
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json"
    }

    # Step 1: Get the current file SHA
    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        sha = response.json().get("sha")
        st.info("✅ Retrieved file SHA from GitHub.")
    else:
        st.error(f"❌ Failed to get file SHA from GitHub: {response.status_code}, {response.text}")
        return

    # Step 2: Read the updated Excel file and encode in base64
    try:
        with open(excel_path, "rb") as f:
            content = f.read()
        encoded_content = base64.b64encode(content).decode()
    except Exception as e:
        st.error(f"❌ Failed to read and encode file: {e}")
        return

    # Step 3: PUT request to update file
    data = {
        "message": "Update Excel with latest testing data",
        "content": encoded_content,
        "sha": sha,
        "branch": "main"  # Or your branch name
    }

    put_response = requests.put(api_url, headers=headers, json=data)
    if put_response.status_code in [200, 201]:
        st.success("✅ Excel file successfully updated on GitHub.")
    else:
        st.error(f"❌ GitHub PUT failed: {put_response.status_code}\n{put_response.text}")
